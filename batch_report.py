"""
批量识别文件夹内所有鸟类照片，生成 Excel + 图片报告
用法: python3 batch_report.py <照片文件夹路径>
"""
import sys
import os
import json
import shutil
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

# 抑制 urllib3 警告
import warnings
warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent))

try:
    from PIL import Image
    from PIL.ExifTags import TAGS
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── EXIF 时间读取 ───────────────────────────────────────────

def get_exif_datetime(image_path: str) -> Optional[datetime]:
    """从 EXIF 读取拍摄时间"""
    if not HAS_PIL:
        return None
    try:
        img = Image.open(image_path)
        exif_data = img._getexif()
        if not exif_data:
            return None
        for tag_id, value in exif_data.items():
            tag = TAGS.get(tag_id, tag_id)
            if tag in ("DateTimeOriginal", "DateTime"):
                return datetime.strptime(value, "%Y:%m:%d %H:%M:%S")
    except Exception:
        pass
    return None


# ── 批量识别 ────────────────────────────────────────────────

def batch_identify(photo_dir: str) -> List[Dict]:
    """批量识别文件夹内所有照片，返回结果列表"""
    from bird_identifier import identify_bird

    photo_dir = Path(photo_dir)
    exts = {".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"}
    photos = sorted([p for p in photo_dir.iterdir() if p.suffix in exts])

    if not photos:
        logger.error(f"文件夹内没有找到照片: {photo_dir}")
        return []

    logger.info(f"找到 {len(photos)} 张照片，开始识别...")
    results = []

    for i, photo_path in enumerate(photos, 1):
        photo_str = str(photo_path)
        shoot_time = get_exif_datetime(photo_str)

        result = identify_bird(photo_str)
        result["file"] = photo_str
        result["filename"] = photo_path.name
        result["shoot_time"] = shoot_time.isoformat() if shoot_time else ""
        result["shoot_time_display"] = shoot_time.strftime("%Y-%m-%d %H:%M") if shoot_time else "未知"
        results.append(result)

        # 进度
        status = result["name_cn"]
        conf = result.get("confidence", 0)
        print(f"\r  [{i:>3}/{len(photos)}] {photo_path.name:<20} → {status:<15} ({conf:.0%})", end="", flush=True)

    print()  # 换行
    return results


# ── Excel 报告 ──────────────────────────────────────────────

def _thumb_for_excel(photo_path: str, size=(120, 90)) -> Optional[str]:
    """生成临时缩略图文件，供 Excel 嵌入"""
    if not HAS_PIL:
        return None
    try:
        import tempfile
        img = Image.open(photo_path)
        img.thumbnail(size, Image.LANCZOS)
        if img.mode in ("RGBA", "P", "CMYK"):
            img = img.convert("RGB")
        tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
        img.save(tmp.name, "JPEG", quality=80)
        return tmp.name
    except Exception as e:
        logger.debug(f"缩略图生成失败 {photo_path}: {e}")
        return None


def generate_excel(results: List[Dict], output_path: str) -> str:
    """生成含缩略图的 Excel 报告"""
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装")
        return ""

    # 分组：已识别 / 未识别
    recognized = [r for r in results if r["name_cn"] not in ("未识别", "待配置识别API")]
    unrecognized = [r for r in results if r["name_cn"] in ("未识别", "待配置识别API")]

    # 按鸟种统计
    species_stats: Dict[str, Dict] = {}
    for r in recognized:
        name = r["name_cn"]
        if name not in species_stats:
            species_stats[name] = {"count": 0, "times": [], "sample_file": r["file"]}
        species_stats[name]["count"] += 1
        if r["shoot_time_display"] != "未知":
            species_stats[name]["times"].append(r["shoot_time_display"])

    wb = openpyxl.Workbook()

    # ── Sheet 1: 汇总 ──
    ws1 = wb.active
    ws1.title = "观鸟汇总"

    # 标题
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = f"观鸟记录报告  {Path(results[0]['file']).parent.name if results else ''}"
    title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2E7D32")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:F2")
    info = ws1["A2"]
    info.value = f"识别总数: {len(results)} 张  |  已识别: {len(recognized)} 张  |  未识别: {len(unrecognized)} 张  |  鸟种数: {len(species_stats)}"
    info.font = Font(name="微软雅黑", size=11, color="1B5E20")
    info.fill = PatternFill("solid", fgColor="C8E6C9")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 24

    # 表头
    headers = ["序号", "缩略图", "鸟种名称", "置信度", "数量", "最早拍摄时间"]
    col_widths = [6, 18, 18, 10, 8, 20]
    header_fill = PatternFill("solid", fgColor="388E3C")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[3].height = 22

    thumb_files = []
    row = 4
    for idx, (name, stat) in enumerate(sorted(species_stats.items(), key=lambda x: -x[1]["count"]), 1):
        ws1.row_dimensions[row].height = 70

        ws1.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=1).border = border

        # 缩略图
        thumb_path = _thumb_for_excel(stat["sample_file"], size=(110, 80))
        if thumb_path:
            thumb_files.append(thumb_path)
            try:
                xl_img = XLImage(thumb_path)
                xl_img.width, xl_img.height = 110, 80
                ws1.add_image(xl_img, f"B{row}")
            except Exception:
                pass
        ws1.cell(row=row, column=2).border = border

        ws1.cell(row=row, column=3, value=name).font = Font(name="微软雅黑", size=12, bold=True)
        ws1.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=3).border = border

        conf_avg = sum(r.get("confidence", 0) for r in recognized if r["name_cn"] == name) / stat["count"]
        ws1.cell(row=row, column=4, value=f"{conf_avg:.0%}").alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=4).border = border

        ws1.cell(row=row, column=5, value=stat["count"]).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=5).border = border

        earliest = min(stat["times"]) if stat["times"] else "未知"
        ws1.cell(row=row, column=6, value=earliest).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=6).border = border

        # 隔行底色
        if idx % 2 == 0:
            for c in range(1, 7):
                ws1.cell(row=row, column=c).fill = PatternFill("solid", fgColor="F1F8E9")

        row += 1

    # ── Sheet 2: 详细记录 ──
    ws2 = wb.create_sheet("详细记录")
    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "每张照片详细识别记录"
    ws2["A1"].font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1565C0")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    detail_headers = ["序号", "文件名", "鸟种名称", "置信度", "拍摄时间", "识别来源", "备注"]
    detail_widths  = [6, 22, 18, 10, 20, 14, 12]
    for col, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(results, 1):
        row_num = i + 2
        note = "待识别新鸟种" if r["name_cn"] in ("未识别", "待配置识别API") else ""
        row_fill = PatternFill("solid", fgColor="FFF3E0") if note else (PatternFill("solid", fgColor="E3F2FD") if i % 2 == 0 else None)

        vals = [i, r["filename"], r["name_cn"], f"{r.get('confidence',0):.0%}",
                r["shoot_time_display"], r.get("source", ""), note]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col == 3 and note:
                cell.font = Font(name="微软雅黑", color="E65100", bold=True)

    # ── Sheet 3: 未识别新鸟种 ──
    ws3 = wb.create_sheet("未识别新鸟种")
    ws3.merge_cells("A1:E1")
    ws3["A1"].value = f"未识别照片（共 {len(unrecognized)} 张）— 可能是训练集以外的新鸟种"
    ws3["A1"].font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="B71C1C")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    un_headers = ["序号", "缩略图", "文件名", "置信度(最高候选)", "拍摄时间"]
    un_widths   = [6, 18, 22, 20, 20]
    for col, (h, w) in enumerate(zip(un_headers, un_widths), 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C62828")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws3.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(unrecognized, 1):
        row_num = i + 2
        ws3.row_dimensions[row_num].height = 70

        ws3.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=row_num, column=1).border = border

        thumb_path = _thumb_for_excel(r["file"], size=(110, 80))
        if thumb_path:
            thumb_files.append(thumb_path)
            try:
                xl_img = XLImage(thumb_path)
                xl_img.width, xl_img.height = 110, 80
                ws3.add_image(xl_img, f"B{row_num}")
            except Exception:
                pass
        ws3.cell(row=row_num, column=2).border = border

        ws3.cell(row=row_num, column=3, value=r["filename"]).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=row_num, column=3).border = border

        # Top1候选（低置信度）
        top5 = r.get("top5", [])
        top1_hint = f"{top5[0][0]} ({top5[0][1]:.0%})" if top5 else f"{r.get('confidence',0):.0%}"
        ws3.cell(row=row_num, column=4, value=top1_hint).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=row_num, column=4).border = border
        ws3.cell(row=row_num, column=4).font = Font(color="B71C1C")

        ws3.cell(row=row_num, column=5, value=r["shoot_time_display"]).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=row_num, column=5).border = border

    wb.save(output_path)
    logger.info(f"Excel 报告已保存: {output_path}")

    # 清理临时缩略图
    for f in thumb_files:
        try:
            os.unlink(f)
        except Exception:
            pass

    return output_path


# ── 图片海报 ────────────────────────────────────────────────

def generate_poster(results: List[Dict], output_path: str) -> str:
    """生成图片海报（每种鸟一格，含缩略图+鸟名+数量）"""
    if not HAS_PIL:
        logger.error("PIL 未安装，无法生成图片报告")
        return ""

    recognized = [r for r in results if r["name_cn"] not in ("未识别", "待配置识别API")]
    unrecognized = [r for r in results if r["name_cn"] in ("未识别", "待配置识别API")]

    # 按鸟种聚合
    species: Dict[str, Dict] = {}
    for r in recognized:
        name = r["name_cn"]
        if name not in species:
            species[name] = {"count": 0, "sample_file": r["file"],
                             "earliest": r["shoot_time_display"]}
        species[name]["count"] += 1
        if r["shoot_time_display"] < species[name]["earliest"] and r["shoot_time_display"] != "未知":
            species[name]["earliest"] = r["shoot_time_display"]

    from PIL import Image, ImageDraw, ImageFont
    import math

    # 布局参数
    CARD_W, CARD_H = 280, 220
    COLS = 4
    ROWS = math.ceil(len(species) / COLS)
    MARGIN = 20
    HEADER_H = 100
    FOOTER_H = 60

    # 未识别区域
    UN_COLS = 6
    UN_ROWS = math.ceil(len(unrecognized) / UN_COLS) if unrecognized else 0
    UN_SECTION_H = (UN_ROWS * 80 + 50) if unrecognized else 0

    total_w = COLS * CARD_W + (COLS + 1) * MARGIN
    total_h = HEADER_H + ROWS * CARD_H + (ROWS + 1) * MARGIN + UN_SECTION_H + FOOTER_H

    canvas = Image.new("RGB", (total_w, total_h), color=(245, 248, 243))
    draw = ImageDraw.Draw(canvas)

    # 字体
    def load_font(size, bold=False):
        candidates = [
            "/System/Library/Fonts/PingFang.ttc",
            "/Library/Fonts/Arial Unicode MS.ttf",
            "/Users/terryoy/Library/Application Support/Ultralytics/Arial.Unicode.ttf",
        ]
        for p in candidates:
            if Path(p).exists():
                try:
                    return ImageFont.truetype(p, size)
                except Exception:
                    pass
        return ImageFont.load_default()

    font_title  = load_font(32, bold=True)
    font_sub    = load_font(16)
    font_name   = load_font(18, bold=True)
    font_small  = load_font(13)

    # 顶部标题栏
    draw.rectangle([0, 0, total_w, HEADER_H], fill=(46, 125, 50))
    folder_name = Path(results[0]["file"]).parent.name if results else ""
    draw.text((total_w // 2, 28), "观鸟记录报告", font=font_title, fill="white", anchor="mm")
    draw.text((total_w // 2, 68),
              f"{folder_name}  |  共 {len(results)} 张  |  {len(species)} 种  |  未识别 {len(unrecognized)} 张",
              font=font_sub, fill="#C8E6C9", anchor="mm")

    # 鸟种卡片
    for idx, (name, stat) in enumerate(sorted(species.items(), key=lambda x: -x[1]["count"])):
        col = idx % COLS
        row = idx // COLS
        x = MARGIN + col * (CARD_W + MARGIN)
        y = HEADER_H + MARGIN + row * (CARD_H + MARGIN)

        # 卡片背景
        draw.rounded_rectangle([x, y, x + CARD_W, y + CARD_H], radius=10, fill="white",
                                outline="#A5D6A7", width=2)

        # 缩略图
        thumb_h = 140
        try:
            thumb = Image.open(stat["sample_file"]).convert("RGB")
            thumb.thumbnail((CARD_W - 4, thumb_h), Image.LANCZOS)
            tw, th = thumb.size
            tx = x + (CARD_W - tw) // 2
            ty = y + 4
            canvas.paste(thumb, (tx, ty))
        except Exception:
            pass

        # 鸟种名
        draw.text((x + CARD_W // 2, y + thumb_h + 14), name,
                  font=font_name, fill="#1B5E20", anchor="mm")
        # 数量 + 时间
        draw.text((x + CARD_W // 2, y + thumb_h + 36),
                  f"{stat['count']} 张  |  {stat['earliest']}",
                  font=font_small, fill="#666666", anchor="mm")

    # 未识别区域
    if unrecognized:
        un_y_start = HEADER_H + ROWS * (CARD_H + MARGIN) + MARGIN * 2
        draw.rectangle([MARGIN, un_y_start, total_w - MARGIN, un_y_start + 36], fill=(183, 28, 28))
        draw.text((total_w // 2, un_y_start + 18),
                  f"未识别照片（{len(unrecognized)} 张）— 可能是新鸟种，建议人工确认",
                  font=font_sub, fill="white", anchor="mm")

        for idx, r in enumerate(unrecognized):
            col = idx % UN_COLS
            row_u = idx // UN_COLS
            ux = MARGIN + col * (total_w - 2 * MARGIN) // UN_COLS
            uy = un_y_start + 44 + row_u * 80
            try:
                thumb = Image.open(r["file"]).convert("RGB")
                thumb.thumbnail((70, 60), Image.LANCZOS)
                canvas.paste(thumb, (ux, uy))
            except Exception:
                pass
            draw.text((ux + 35, uy + 64), Path(r["file"]).name[:12],
                      font=ImageFont.load_default(), fill="#555555", anchor="mm")

    # 底部
    draw.rectangle([0, total_h - FOOTER_H, total_w, total_h], fill=(200, 230, 201))
    draw.text((total_w // 2, total_h - FOOTER_H // 2),
              f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  本地 YOLOv8 模型识别",
              font=font_small, fill="#2E7D32", anchor="mm")

    canvas.save(output_path, "JPEG", quality=92)
    logger.info(f"图片报告已保存: {output_path}")
    return output_path


# ── 主程序 ──────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 batch_report.py <照片文件夹路径>")
        sys.exit(1)

    photo_dir = sys.argv[1]
    folder_name = Path(photo_dir).name
    output_dir = Path(photo_dir).parent
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"\n{'='*60}")
    print(f"  观鸟助手 — 批量识别报告")
    print(f"  文件夹: {photo_dir}")
    print(f"{'='*60}\n")

    # 批量识别
    results = batch_identify(photo_dir)
    if not results:
        sys.exit(1)

    # 统计
    recognized = [r for r in results if r["name_cn"] not in ("未识别", "待配置识别API")]
    unrecognized = [r for r in results if r["name_cn"] in ("未识别", "待配置识别API")]
    species_set = set(r["name_cn"] for r in recognized)

    print(f"\n{'='*60}")
    print(f"  识别完成！")
    print(f"  总计: {len(results)} 张  已识别: {len(recognized)} 张  未识别: {len(unrecognized)} 张")
    print(f"  发现鸟种: {len(species_set)} 种 — {', '.join(sorted(species_set))}")
    print(f"{'='*60}\n")

    # 保存 JSON（方便调试）
    json_path = str(output_dir / f"识别结果_{folder_name}_{ts}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # 生成 Excel
    excel_path = str(output_dir / f"观鸟报告_{folder_name}_{ts}.xlsx")
    generate_excel(results, excel_path)

    # 生成图片海报
    poster_path = str(output_dir / f"观鸟海报_{folder_name}_{ts}.jpg")
    generate_poster(results, poster_path)

    print(f"\n输出文件：")
    print(f"  Excel 报告: {excel_path}")
    print(f"  图片 报告: {poster_path}")
    print(f"  识别 JSON: {json_path}")
