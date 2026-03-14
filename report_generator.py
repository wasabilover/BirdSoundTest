"""
观鸟报告生成模块
生成 Excel 报告和图片海报报告（可发朋友圈/分享好友）
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

from config import (
    REPORT_OUTPUT_DIR,
    REPORT_COLS,
    THUMBNAIL_SIZE,
    REPORT_TITLE,
    OBSERVER_NAME,
)

logger = logging.getLogger(__name__)

# 颜色配置
COLOR_BG = "#1A2332"          # 深蓝背景
COLOR_CARD = "#243447"        # 卡片背景
COLOR_ACCENT = "#4EC9B0"      # 青绿强调色
COLOR_TITLE = "#FFFFFF"       # 标题白
COLOR_SUBTITLE = "#A8C7E8"    # 副标题浅蓝
COLOR_TEXT = "#D4E9F7"        # 正文浅色
COLOR_GOLD = "#FFD700"        # 金色（新物种标记）


def _hex_to_rgb(hex_color: str):
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def generate_excel_report(
    session_records: List[Dict],
    output_path: Optional[str] = None,
) -> Optional[str]:
    """生成 Excel 观鸟报告"""
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装，无法生成 Excel 报告")
        return None

    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(REPORT_OUTPUT_DIR, f"观鸟报告_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    ws_detail = wb.create_sheet("详细记录")

    # ── 汇总页 ──────────────────────────────────────────
    _style_excel_summary(ws_summary, session_records)

    # ── 详细记录页 ──────────────────────────────────────
    _style_excel_detail(ws_detail, session_records)

    wb.save(output_path)
    logger.info(f"Excel 报告已生成: {output_path}")
    return output_path


def _make_header_fill(color_hex: str):
    return PatternFill("solid", fgColor=color_hex.lstrip("#"))


def _border_thin():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_excel_summary(ws, records: List[Dict]):
    """汇总页样式"""
    # 统计物种
    species_count = defaultdict(lambda: {"count": 0, "sci": "", "dates": [], "photos": []})
    for r in records:
        name = r.get("bird_cn", "未识别")
        species_count[name]["count"] += 1
        species_count[name]["sci"] = r.get("bird_sci", "")
        species_count[name]["dates"].append(r.get("date", ""))
        species_count[name]["photos"].append(r.get("file", ""))

    session_date = datetime.now().strftime("%Y年%m月%d日")

    # 标题行
    ws.merge_cells("A1:F1")
    ws["A1"] = f"🦅 {REPORT_TITLE}"
    ws["A1"].font = Font(name="微软雅黑", size=18, bold=True, color="1A5276")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = _make_header_fill("D6EAF8")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"] = f"观鸟者：{OBSERVER_NAME}    日期：{session_date}    共识别 {len(records)} 张照片    发现 {len(species_count)} 种鸟类"
    ws["A2"].font = Font(name="微软雅黑", size=11, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # 表头
    headers = ["#", "中文名", "学名", "数量（张）", "首次拍摄时间", "置信度"]
    col_widths = [5, 18, 28, 12, 20, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = _make_header_fill("1A5276")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    # 数据行（按数量排序）
    sorted_species = sorted(species_count.items(), key=lambda x: -x[1]["count"])
    for row_idx, (name, info) in enumerate(sorted_species, 4):
        dates = sorted([d for d in info["dates"] if d])
        first_date = dates[0][:10] if dates else ""
        # 最高置信度（需从 records 查）
        max_conf = max(
            (r.get("confidence", 0) for r in records if r.get("bird_cn") == name),
            default=0
        )

        row_fill = _make_header_fill("EBF5FB") if row_idx % 2 == 0 else _make_header_fill("FFFFFF")
        data = [row_idx - 3, name, info["sci"], info["count"], first_date, f"{max_conf:.0%}"]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left", vertical="center")
            cell.fill = row_fill
            cell.border = _border_thin()
        ws.row_dimensions[row_idx].height = 20

    # 冻结标题行
    ws.freeze_panes = "A4"


def _style_excel_detail(ws, records: List[Dict]):
    """详细记录页样式"""
    headers = ["序号", "文件名", "中文名", "学名", "拍摄时间", "置信度", "识别来源", "归档路径"]
    col_widths = [6, 22, 14, 24, 18, 10, 14, 50]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = _make_header_fill("2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, r in enumerate(records, 2):
        row_fill = _make_header_fill("F0F3F4") if row_idx % 2 == 0 else _make_header_fill("FFFFFF")
        data = [
            row_idx - 1,
            r.get("original_name", ""),
            r.get("bird_cn", ""),
            r.get("bird_sci", ""),
            r.get("date", "")[:16] if r.get("date") else "",
            f"{r.get('confidence', 0):.0%}",
            r.get("source", ""),
            r.get("file", ""),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=9)
            cell.alignment = Alignment(horizontal="center" if col_idx not in (2, 4, 8) else "left",
                                       vertical="center")
            cell.fill = row_fill
            cell.border = _border_thin()
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(records)+1}"


# ─────────────────────────────────────────────────────────────
# 图片报告（适合发朋友圈）
# ─────────────────────────────────────────────────────────────

def _load_font(size: int, bold: bool = False):
    """加载中文字体，尝试系统字体"""
    font_candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/Library/Fonts/Arial Unicode MS.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
    ]
    for fp in font_candidates:
        if os.path.exists(fp):
            try:
                from PIL import ImageFont
                return ImageFont.truetype(fp, size)
            except Exception:
                continue
    from PIL import ImageFont
    return ImageFont.load_default()


def _draw_rounded_rect(draw, xy, radius, fill):
    """画圆角矩形"""
    x0, y0, x1, y1 = xy
    draw.rounded_rectangle([x0, y0, x1, y1], radius=radius, fill=fill)


def _make_thumbnail(photo_path: str, size: tuple) -> Optional["Image.Image"]:
    """生成缩略图，失败返回占位图"""
    if not HAS_PIL:
        return None
    try:
        img = Image.open(photo_path)
        img.thumbnail(size, Image.LANCZOS)
        # 居中裁剪到固定尺寸
        bg = Image.new("RGB", size, (30, 50, 70))
        offset = ((size[0] - img.width) // 2, (size[1] - img.height) // 2)
        bg.paste(img, offset)
        return bg
    except Exception:
        # 占位灰色图
        ph = Image.new("RGB", size, (50, 70, 90))
        draw = ImageDraw.Draw(ph)
        font = _load_font(18)
        draw.text((size[0]//2, size[1]//2), "无图片", fill=(150, 170, 190),
                  font=font, anchor="mm")
        return ph


def generate_image_report(
    session_records: List[Dict],
    output_path: Optional[str] = None,
) -> Optional[str]:
    """
    生成美观的图片海报式报告（适合发朋友圈）
    """
    if not HAS_PIL:
        logger.error("Pillow 未安装，无法生成图片报告")
        return None

    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(REPORT_OUTPUT_DIR, f"观鸟海报_{ts}.jpg")

    # 统计物种，取每种第一张照片作为代表
    species_map = {}  # name_cn -> {info..., representative_photo}
    for r in session_records:
        name = r.get("bird_cn", "未识别")
        if name not in species_map:
            species_map[name] = {
                "sci": r.get("bird_sci", ""),
                "count": 0,
                "confidence": r.get("confidence", 0),
                "photo": r.get("file", ""),
                "date": r.get("date", ""),
            }
        species_map[name]["count"] += 1

    species_list = sorted(species_map.items(), key=lambda x: -x[1]["count"])
    num_species = len(species_list)

    # ── 布局计算 ──────────────────────────────────────────
    COLS = min(REPORT_COLS, num_species) if num_species > 0 else 1
    ROWS = (num_species + COLS - 1) // COLS

    CARD_W, CARD_H = THUMBNAIL_SIZE[0] + 20, THUMBNAIL_SIZE[1] + 80
    MARGIN = 30
    HEADER_H = 160
    FOOTER_H = 80

    TOTAL_W = COLS * CARD_W + (COLS + 1) * MARGIN
    TOTAL_H = HEADER_H + ROWS * CARD_H + (ROWS + 1) * MARGIN + FOOTER_H
    TOTAL_W = max(TOTAL_W, 900)

    canvas = Image.new("RGB", (TOTAL_W, TOTAL_H), _hex_to_rgb(COLOR_BG))
    draw = ImageDraw.Draw(canvas)

    # ── 顶部渐变标题区 ────────────────────────────────────
    for y in range(HEADER_H):
        ratio = y / HEADER_H
        r = int(26 + (40 - 26) * ratio)
        g = int(35 + (60 - 35) * ratio)
        b = int(50 + (80 - 50) * ratio)
        draw.line([(0, y), (TOTAL_W, y)], fill=(r, g, b))

    # 标题文字
    font_title = _load_font(42, bold=True)
    font_sub = _load_font(20)
    font_small = _load_font(16)

    session_date = datetime.now().strftime("%Y年%m月%d日")
    draw.text(
        (TOTAL_W // 2, 52),
        f"🦅 {REPORT_TITLE}",
        fill=_hex_to_rgb(COLOR_TITLE),
        font=font_title,
        anchor="mm",
    )
    draw.text(
        (TOTAL_W // 2, 100),
        f"{session_date}  ·  {OBSERVER_NAME}  ·  发现 {num_species} 种  共 {len(session_records)} 张",
        fill=_hex_to_rgb(COLOR_SUBTITLE),
        font=font_sub,
        anchor="mm",
    )
    # 分隔线
    draw.line([(MARGIN, HEADER_H - 20), (TOTAL_W - MARGIN, HEADER_H - 20)],
              fill=_hex_to_rgb(COLOR_ACCENT), width=2)

    # ── 鸟类卡片 ──────────────────────────────────────────
    font_name = _load_font(18, bold=True)
    font_sci = _load_font(13)
    font_count = _load_font(14)

    for idx, (name, info) in enumerate(species_list):
        col = idx % COLS
        row = idx // COLS

        card_x = MARGIN + col * (CARD_W + MARGIN)
        card_y = HEADER_H + MARGIN + row * (CARD_H + MARGIN)

        # 卡片背景（圆角）
        _draw_rounded_rect(
            draw,
            (card_x, card_y, card_x + CARD_W, card_y + CARD_H),
            radius=12,
            fill=_hex_to_rgb(COLOR_CARD),
        )

        # 照片缩略图
        thumb = _make_thumbnail(info["photo"], THUMBNAIL_SIZE)
        if thumb:
            # 给缩略图加圆角遮罩
            mask = Image.new("L", THUMBNAIL_SIZE, 0)
            mask_draw = ImageDraw.Draw(mask)
            mask_draw.rounded_rectangle([0, 0, THUMBNAIL_SIZE[0]-1, THUMBNAIL_SIZE[1]-1],
                                        radius=8, fill=255)
            thumb_x = card_x + 10
            thumb_y = card_y + 8
            canvas.paste(thumb, (thumb_x, thumb_y), mask)

        # 鸟名
        text_y = card_y + THUMBNAIL_SIZE[1] + 16
        draw.text(
            (card_x + CARD_W // 2, text_y),
            name,
            fill=_hex_to_rgb(COLOR_TITLE),
            font=font_name,
            anchor="mm",
        )

        # 学名
        sci_name = info["sci"]
        if sci_name and sci_name != "Unknown":
            draw.text(
                (card_x + CARD_W // 2, text_y + 24),
                sci_name,
                fill=_hex_to_rgb(COLOR_SUBTITLE),
                font=font_sci,
                anchor="mm",
            )

        # 数量标签
        count_text = f"× {info['count']}"
        badge_w = 50
        badge_x = card_x + CARD_W - badge_w - 8
        badge_y = card_y + 8
        _draw_rounded_rect(
            draw,
            (badge_x, badge_y, badge_x + badge_w, badge_y + 24),
            radius=10,
            fill=_hex_to_rgb(COLOR_ACCENT),
        )
        draw.text(
            (badge_x + badge_w // 2, badge_y + 12),
            count_text,
            fill=(20, 20, 30),
            font=font_count,
            anchor="mm",
        )

        # 置信度（小标）
        conf = info["confidence"]
        if conf > 0:
            conf_text = f"{conf:.0%}"
            draw.text(
                (card_x + 14, card_y + THUMBNAIL_SIZE[1] + 16),
                conf_text,
                fill=_hex_to_rgb(COLOR_GOLD),
                font=font_small,
                anchor="lm",
            )

    # ── 底部 ──────────────────────────────────────────────
    footer_y = TOTAL_H - FOOTER_H + 20
    draw.line([(MARGIN, footer_y), (TOTAL_W - MARGIN, footer_y)],
              fill=(60, 80, 100), width=1)
    draw.text(
        (TOTAL_W // 2, footer_y + 30),
        "Generated by 观鸟助手 · Bird Watcher Assistant",
        fill=(80, 100, 130),
        font=font_small,
        anchor="mm",
    )

    # 轻微锐化
    canvas = canvas.filter(ImageFilter.SHARPEN)
    canvas.save(output_path, "JPEG", quality=95)
    logger.info(f"图片报告已生成: {output_path}")
    return output_path


def generate_reports(
    session_records: List[Dict],
    session_name: Optional[str] = None,
) -> Dict[str, str]:
    """生成所有报告"""
    results = {}
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = session_name or ts

    excel_path = os.path.join(REPORT_OUTPUT_DIR, f"观鸟报告_{prefix}.xlsx")
    image_path = os.path.join(REPORT_OUTPUT_DIR, f"观鸟海报_{prefix}.jpg")

    xlsx = generate_excel_report(session_records, excel_path)
    img = generate_image_report(session_records, image_path)

    if xlsx:
        results["excel"] = xlsx
    if img:
        results["image"] = img

    return results
